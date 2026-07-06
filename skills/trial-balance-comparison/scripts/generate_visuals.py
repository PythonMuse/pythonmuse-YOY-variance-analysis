"""
generate_visuals.py - CodeCritters Inc. Trial Balance Comparison
PythonMuse LLC | skills/trial-balance-comparison

Reads data/trial_balance_2025.csv and data/trial_balance_2026.csv, runs a
Year-over-Year and Month-over-Month variance pass, flags accounts that clear
BOTH a dollar and a percent materiality threshold, and writes:
  - outputs/visuals/01_net_income_trend.png
  - outputs/visuals/02_yoy_variance_by_account.png
  - outputs/visuals/03_flagged_accounts_trend.png
  - outputs/excel/CodeCritters_TB_Comparison_Report.xlsx

Never modifies the source CSVs (CLAUDE.md, Rule 9). Every cell in the
workbook that looks like a calculation IS a calculation - Excel formulas
referencing the raw data sheets, not Python-computed values pasted in
(CLAUDE.md, Rule 8).

Usage:
    python3 skills/trial-balance-comparison/scripts/generate_visuals.py
"""

import os

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG - the only knobs you should need to touch
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", ".."))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
VISUALS_DIR = os.path.join(PROJECT_ROOT, "outputs", "visuals")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "outputs", "excel")
os.makedirs(VISUALS_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

PRIOR_CSV = os.path.join(DATA_DIR, "trial_balance_2025.csv")
CURRENT_CSV = os.path.join(DATA_DIR, "trial_balance_2026.csv")
REPORT_PATH = os.path.join(OUTPUT_DIR, "CodeCritters_TB_Comparison_Report.xlsx")

MONTHS = ["Jan", "Feb", "Mar", "Apr"]

# >>> HUMAN INPUT NEEDED: adjust materiality thresholds here, not in a
# one-off chat answer. These also get written into the Dashboard sheet
# so a reviewer can see exactly what the flags mean without opening this file.
DOLLAR_THRESHOLD_YOY = 2500
PCT_THRESHOLD_YOY = 0.15
DOLLAR_THRESHOLD_MOM = 2500
PCT_THRESHOLD_MOM = 0.20

# Point-in-time balance vs. period-activity classification (CLAUDE.md, Rule 1)
FLOW_TYPES = {"Revenue", "COGS", "Expense"}

# Contra accounts carry a natural balance opposite their parent category and
# must be SUBTRACTED, not summed in, when reconciling or totaling by category.
CONTRA_ASSETS = {"Accumulated Depreciation"}
CONTRA_EQUITY = {"Owner Draws"}

# PythonMuse brand palette (see skills/pythonmuse-visual-branding)
DEEP_NAVY = "#002030"
PANEL_NAVY = "#0A3040"
PRIMARY_TEAL = "#1E8890"
CIRCUIT_TEAL = "#2A6870"
GOLD_AMBER = "#F8D038"
WARM_CREAM = "#F8E098"
NEAR_WHITE_CYAN = "#F0F8F8"
WHITE = "#FFFFFF"
LIGHT_GRAY = "#F5F5F5"

FOOTER_TEXT = "PythonMuse LLC  |  github.com/PythonMuse/ai-ledger"


def style_ax(ax, title, ylabel=None):
    """White plot area, dark navy text - brand contrast rule #1."""
    ax.set_facecolor(LIGHT_GRAY)
    ax.set_title(title, fontsize=15, fontweight="bold", color=DEEP_NAVY, pad=14)
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=11, color=DEEP_NAVY)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.tick_params(colors=DEEP_NAVY, labelsize=10)


# ============================================================
# LOAD + RECONCILE
# ============================================================

def load_and_reconcile(path, label):
    df = pd.read_csv(path)
    for m in MONTHS:
        assets = df.loc[(df["Type"] == "Asset") & (~df["Account"].isin(CONTRA_ASSETS)), m].sum()
        assets -= df.loc[df["Account"].isin(CONTRA_ASSETS), m].sum()
        equity = df.loc[(df["Type"] == "Equity") & (~df["Account"].isin(CONTRA_EQUITY)), m].sum()
        equity -= df.loc[df["Account"].isin(CONTRA_EQUITY), m].sum()
        liab = df.loc[df["Type"] == "Liability", m].sum()
        liab_eq = liab + equity
        diff = round(assets - liab_eq, 2)
        if diff != 0:
            raise SystemExit(
                f"[STOP] {label} does not reconcile in {m}: "
                f"Assets {assets} != Liabilities + Equity {liab_eq} (diff {diff}). "
                f"Fix source data before running variance analysis (CLAUDE.md, Rule 5)."
            )
    print(f"[OK] {label} reconciles: Assets = Liabilities + Equity, every month.")
    return df


prior_df = load_and_reconcile(PRIOR_CSV, "trial_balance_2025.csv")
curr_df = load_and_reconcile(CURRENT_CSV, "trial_balance_2026.csv")

assert list(prior_df["Account Code"]) == list(curr_df["Account Code"]), \
    "Account lists differ between the two files - cannot compare row-by-row."

# ============================================================
# YoY VARIANCE PASS - same month, current year vs. prior year
# ============================================================

yoy_rows = []
for i in range(len(prior_df)):
    code, acct, atype = prior_df.loc[i, ["Account Code", "Account", "Type"]]
    for m in MONTHS:
        prior_val = float(prior_df.loc[i, m])
        curr_val = float(curr_df.loc[i, m])
        dollar_var = curr_val - prior_val
        pct_var = (dollar_var / abs(prior_val)) if prior_val != 0 else None
        flagged = abs(dollar_var) > DOLLAR_THRESHOLD_YOY and pct_var is not None and abs(pct_var) > PCT_THRESHOLD_YOY
        yoy_rows.append(dict(code=code, account=acct, type=atype, month=m,
                              prior=prior_val, current=curr_val,
                              dollar_var=dollar_var, pct_var=pct_var, flagged=flagged))

yoy_df = pd.DataFrame(yoy_rows)
yoy_flagged = yoy_df[yoy_df["flagged"]]

# ============================================================
# MoM VARIANCE PASS - sequential months, current year only
# ============================================================

mom_rows = []
for i in range(len(curr_df)):
    code, acct, atype = curr_df.loc[i, ["Account Code", "Account", "Type"]]
    for m1, m2 in zip(MONTHS[:-1], MONTHS[1:]):
        v1 = float(curr_df.loc[i, m1])
        v2 = float(curr_df.loc[i, m2])
        dollar_var = v2 - v1
        pct_var = (dollar_var / abs(v1)) if v1 != 0 else None
        flagged = abs(dollar_var) > DOLLAR_THRESHOLD_MOM and pct_var is not None and abs(pct_var) > PCT_THRESHOLD_MOM
        mom_rows.append(dict(code=code, account=acct, type=atype, from_month=m1, to_month=m2,
                              v1=v1, v2=v2, dollar_var=dollar_var, pct_var=pct_var, flagged=flagged))

mom_df = pd.DataFrame(mom_rows)
mom_flagged = mom_df[mom_df["flagged"]]

# Rule 6: is a MoM flag a recurring seasonal pattern also present in the prior year?
def is_recurring(row):
    i = prior_df.index[prior_df["Account Code"] == row["code"]][0]
    v1 = float(prior_df.loc[i, row["from_month"]])
    v2 = float(prior_df.loc[i, row["to_month"]])
    prior_dollar_var = v2 - v1
    # same direction and both cross the dollar bar in the prior year too
    same_direction = (prior_dollar_var > 0) == (row["dollar_var"] > 0)
    return same_direction and abs(prior_dollar_var) > DOLLAR_THRESHOLD_MOM

mom_df["recurring"] = mom_df.apply(is_recurring, axis=1)
mom_flagged = mom_df[mom_df["flagged"]].copy()

print(f"\nYoY flags: {len(yoy_flagged)}   MoM flags: {len(mom_flagged)}")

# ============================================================
# CHART 1 - Net Income Trend, 2025 vs 2026
# ============================================================

def net_income_by_month(df):
    ni = []
    for m in MONTHS:
        rev = df.loc[df["Type"] == "Revenue", m].sum()
        cogs = df.loc[df["Type"] == "COGS", m].sum()
        exp = df.loc[df["Type"] == "Expense", m].sum()
        ni.append(rev - cogs - exp)
    return ni

ni_2025 = net_income_by_month(prior_df)
ni_2026 = net_income_by_month(curr_df)

fig, ax = plt.subplots(figsize=(9, 5))
fig.patch.set_facecolor(WHITE)
ax.plot(MONTHS, ni_2025, color=PRIMARY_TEAL, marker="o", linewidth=2.5, markersize=8, label="2025")
ax.plot(MONTHS, ni_2026, color=GOLD_AMBER, marker="D", linewidth=2.5, markersize=8, label="2026")
y_span = max(max(ni_2025), max(ni_2026)) - min(min(ni_2025), min(ni_2026))
offset = y_span * 0.06
for i, (v25, v26) in enumerate(zip(ni_2025, ni_2026)):
    top_val, bot_val = (v26, v25) if v26 >= v25 else (v25, v26)
    top_color = GOLD_AMBER if v26 >= v25 else PRIMARY_TEAL
    bot_color = PRIMARY_TEAL if v26 >= v25 else GOLD_AMBER
    top_weight = "bold" if v26 >= v25 else "normal"
    bot_weight = "normal" if v26 >= v25 else "bold"
    ax.text(i, top_val + offset, f"${top_val:,.0f}", ha="center", fontsize=9,
            color=DEEP_NAVY if top_color == GOLD_AMBER else top_color, fontweight=top_weight)
    ax.text(i, bot_val - offset, f"${bot_val:,.0f}", ha="center", fontsize=9,
            color=DEEP_NAVY if bot_color == GOLD_AMBER else bot_color, fontweight=bot_weight)
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
style_ax(ax, "Net Income by Month - CodeCritters Inc.", "Net Income ($)")
ax.legend(fontsize=11, loc="lower right", framealpha=0.9)
ax.set_ylim(min(min(ni_2025), min(ni_2026)) - offset * 2.2, max(max(ni_2025), max(ni_2026)) + offset * 2.2)
plt.tight_layout(rect=[0, 0.04, 1, 1])
fig.text(0.5, 0.01, FOOTER_TEXT, ha="center", fontsize=8, color=DEEP_NAVY, alpha=0.55)
plt.savefig(os.path.join(VISUALS_DIR, "01_net_income_trend.png"), dpi=180)
plt.close()
print("[OK] 01_net_income_trend.png")

# ============================================================
# CHART 2 - YoY Variance by Account (period totals, trap-safe)
# ============================================================
# Balance-sheet accounts: compare Apr ending balance (point-in-time).
# P&L accounts: compare summed Jan-Apr activity. Never the other way
# around (CLAUDE.md, Rule 1 / README "First Trap").

acct_var = []
for i in range(len(prior_df)):
    code, acct, atype = prior_df.loc[i, ["Account Code", "Account", "Type"]]
    if atype in FLOW_TYPES:
        prior_val = prior_df.loc[i, MONTHS].sum()
        curr_val = curr_df.loc[i, MONTHS].sum()
    else:
        prior_val = prior_df.loc[i, "Apr"]
        curr_val = curr_df.loc[i, "Apr"]
    var = curr_val - prior_val
    flagged_any = bool(yoy_df[(yoy_df["code"] == code) & (yoy_df["flagged"])].shape[0])
    acct_var.append((acct, var, flagged_any))

acct_var_df = pd.DataFrame(acct_var, columns=["account", "var", "flagged"])
top = acct_var_df.reindex(acct_var_df["var"].abs().sort_values(ascending=False).index).head(15)
top = top.iloc[::-1]  # largest at top of horizontal bar chart

fig, ax = plt.subplots(figsize=(9, 7))
fig.patch.set_facecolor(WHITE)
colors = [GOLD_AMBER if f else PRIMARY_TEAL for f in top["flagged"]]
bars = ax.barh(top["account"], top["var"], color=colors, edgecolor="white", linewidth=0.5)
ax.axvline(0, color="#999999", linewidth=0.8)
for bar, v in zip(bars, top["var"]):
    ax.text(v + (500 if v >= 0 else -500), bar.get_y() + bar.get_height() / 2,
            f"${v:,.0f}", va="center", ha="left" if v >= 0 else "right",
            fontsize=9, color=DEEP_NAVY)
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
style_ax(ax, "YoY Variance by Account (Jan-Apr), Top 15 by Size", "")
ax.tick_params(axis="y", labelsize=10)
plt.tight_layout(rect=[0, 0.04, 1, 1])
fig.text(0.5, 0.01, f"{FOOTER_TEXT}   |   Gold = crosses both YoY thresholds in at least one month",
          ha="center", fontsize=8, color=DEEP_NAVY, alpha=0.55)
plt.savefig(os.path.join(VISUALS_DIR, "02_yoy_variance_by_account.png"), dpi=180)
plt.close()
print("[OK] 02_yoy_variance_by_account.png")

# ============================================================
# CHART 3 - Flagged Accounts Worth a Human's Five Minutes
# ============================================================

spotlight = ["COGS - Materials", "Marketing & Advertising", "Accounts Receivable", "Professional Fees"]

fig, axes = plt.subplots(2, 2, figsize=(11, 8))
fig.patch.set_facecolor(WHITE)
for ax, acct in zip(axes.flat, spotlight):
    p_row = prior_df[prior_df["Account"] == acct].iloc[0]
    c_row = curr_df[curr_df["Account"] == acct].iloc[0]
    ax.plot(MONTHS, [p_row[m] for m in MONTHS], color=PRIMARY_TEAL, marker="o", linewidth=2, label="2025")
    ax.plot(MONTHS, [c_row[m] for m in MONTHS], color=GOLD_AMBER, marker="D", linewidth=2, label="2026")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    style_ax(ax, acct)
    ax.legend(fontsize=9, loc="best", framealpha=0.9)
fig.suptitle("Flagged Accounts Worth a Human's Five Minutes", fontsize=16, fontweight="bold", color=DEEP_NAVY, y=0.99)
plt.tight_layout(rect=[0, 0.03, 1, 0.96])
fig.text(0.5, 0.005, FOOTER_TEXT, ha="center", fontsize=8, color=DEEP_NAVY, alpha=0.55)
plt.savefig(os.path.join(VISUALS_DIR, "03_flagged_accounts_trend.png"), dpi=180)
plt.close()
print("[OK] 03_flagged_accounts_trend.png")

print(f"\nAll charts saved to: {VISUALS_DIR}")

# ============================================================
# EXCEL REPORT - every calculated cell is a formula, not a value
# (CLAUDE.md, Rule 8)
# ============================================================

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor=PANEL_NAVY.lstrip("#"))
HEADER_FONT = Font(color="FFFFFF", bold=True)
FLAG_FILL = PatternFill("solid", fgColor=GOLD_AMBER.lstrip("#"))
TITLE_FONT = Font(bold=True, size=14, color=DEEP_NAVY.lstrip("#"))


def write_header(ws, row, headers, start_col=1):
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---- TB_2025 / TB_2026 (raw source, read-only in spirit - Rule 9) ----

def write_tb_sheet(ws, df, year_label):
    ws.title = ws.title
    ws.cell(row=1, column=1, value=f"Trial Balance - {year_label} (source data, do not edit)").font = TITLE_FONT
    write_header(ws, 2, ["Account Code", "Account", "Type"] + MONTHS)
    for i in range(len(df)):
        r = 3 + i
        ws.cell(row=r, column=1, value=df.loc[i, "Account Code"])
        ws.cell(row=r, column=2, value=df.loc[i, "Account"])
        ws.cell(row=r, column=3, value=df.loc[i, "Type"])
        for j, m in enumerate(MONTHS):
            ws.cell(row=r, column=4 + j, value=float(df.loc[i, m])).number_format = "#,##0.00"
    autosize(ws, {"A": 14, "B": 26, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12})


ws_2025 = wb.active
ws_2025.title = "TB_2025"
write_tb_sheet(ws_2025, prior_df, "2025 (Prior Year)")

ws_2026 = wb.create_sheet("TB_2026")
write_tb_sheet(ws_2026, curr_df, "2026 (Current Year)")

TB_DATA_START_ROW = 3  # row where account data starts on TB_2025 / TB_2026
MONTH_COL = {m: get_column_letter(4 + j) for j, m in enumerate(MONTHS)}  # Jan=D ... Apr=G

# ---- Dashboard ----

dash = wb.create_sheet("Dashboard", 0)
dash.cell(row=1, column=1, value="CodeCritters Inc. - Trial Balance Comparison Dashboard").font = Font(bold=True, size=16, color=DEEP_NAVY.lstrip("#"))
dash.cell(row=2, column=1, value="PythonMuse LLC | Article 33: Two Trial Balances Walk Into a Spreadsheet").font = Font(italic=True, size=10)

dash.cell(row=4, column=1, value="Materiality Thresholds").font = Font(bold=True, size=12, color=PANEL_NAVY.lstrip("#"))
write_header(dash, 5, ["Comparison", "$ Threshold", "% Threshold"])
dash.cell(row=6, column=1, value="Year-over-Year")
dash.cell(row=6, column=2, value=DOLLAR_THRESHOLD_YOY)
dash.cell(row=6, column=3, value=PCT_THRESHOLD_YOY).number_format = "0%"
dash.cell(row=7, column=1, value="Month-over-Month")
dash.cell(row=7, column=2, value=DOLLAR_THRESHOLD_MOM)
dash.cell(row=7, column=3, value=PCT_THRESHOLD_MOM).number_format = "0%"
# Named thresholds cells for downstream formulas
YOY_DOLLAR_CELL, YOY_PCT_CELL = "Dashboard!$B$6", "Dashboard!$C$6"
MOM_DOLLAR_CELL, MOM_PCT_CELL = "Dashboard!$B$7", "Dashboard!$C$7"

dash.cell(row=9, column=1, value="Flag Counts").font = Font(bold=True, size=12, color=PANEL_NAVY.lstrip("#"))
dash.cell(row=10, column=1, value="YoY flags")
dash.cell(row=10, column=2, value="=COUNTIF(YoY_Variance!I3:I10000,\"FLAG\")")
dash.cell(row=11, column=1, value="MoM flags")
dash.cell(row=11, column=2, value="=COUNTIF(MoM_Variance!J3:J10000,\"FLAG\")")

dash.cell(row=13, column=1, value="Reconciliation Check (Assets = Liabilities + Equity, net of contras)").font = Font(bold=True, size=12, color=PANEL_NAVY.lstrip("#"))
write_header(dash, 14, ["Year", "Month", "Assets (net)", "Liabilities + Equity (net)", "Diff", "Status"])
recon_row = 15
last_tb_row = TB_DATA_START_ROW + len(prior_df) - 1
for year_label, sheet in [("2025", "TB_2025"), ("2026", "TB_2026")]:
    for m in MONTHS:
        col = MONTH_COL[m]
        rng = f"'{sheet}'!${col}${TB_DATA_START_ROW}:${col}${last_tb_row}"
        type_rng = f"'{sheet}'!$C${TB_DATA_START_ROW}:$C${last_tb_row}"
        acct_rng = f"'{sheet}'!$B${TB_DATA_START_ROW}:$B${last_tb_row}"
        dash.cell(row=recon_row, column=1, value=year_label)
        dash.cell(row=recon_row, column=2, value=m)
        # Assets net of contra-assets (e.g. Accumulated Depreciation)
        assets_formula = f"=SUMIF({type_rng},\"Asset\",{rng})-SUMIF({acct_rng},\"Accumulated Depreciation\",{rng})*2"
        dash.cell(row=recon_row, column=3, value=assets_formula)
        # Liabilities + Equity, net of contra-equity (e.g. Owner Draws)
        liab_eq_formula = (
            f"=SUMIF({type_rng},\"Liability\",{rng})+SUMIF({type_rng},\"Equity\",{rng})"
            f"-SUMIF({acct_rng},\"Owner Draws\",{rng})*2"
        )
        dash.cell(row=recon_row, column=4, value=liab_eq_formula)
        dash.cell(row=recon_row, column=5, value=f"=C{recon_row}-D{recon_row}")
        dash.cell(row=recon_row, column=6, value=f"=IF(ROUND(E{recon_row},2)=0,\"OK\",\"MISMATCH\")")
        recon_row += 1

autosize(dash, {"A": 20, "B": 14, "C": 18, "D": 20, "E": 14, "F": 12})

# ---- YoY_Variance ----

yoy_ws = wb.create_sheet("YoY_Variance")
yoy_ws.cell(row=1, column=1, value="Year-over-Year Variance - same month, 2026 vs. 2025").font = TITLE_FONT
write_header(yoy_ws, 2, ["Account Code", "Account", "Type", "Month", "Prior (2025)", "Current (2026)", "$ Var", "% Var", "Flag"])

row = 3
for i in range(len(prior_df)):
    tb_row = TB_DATA_START_ROW + i
    for m in MONTHS:
        col = MONTH_COL[m]
        yoy_ws.cell(row=row, column=1, value=f"='TB_2025'!A{tb_row}")
        yoy_ws.cell(row=row, column=2, value=f"='TB_2025'!B{tb_row}")
        yoy_ws.cell(row=row, column=3, value=f"='TB_2025'!C{tb_row}")
        yoy_ws.cell(row=row, column=4, value=m)
        yoy_ws.cell(row=row, column=5, value=f"='TB_2025'!{col}{tb_row}").number_format = "#,##0.00"
        yoy_ws.cell(row=row, column=6, value=f"='TB_2026'!{col}{tb_row}").number_format = "#,##0.00"
        yoy_ws.cell(row=row, column=7, value=f"=F{row}-E{row}").number_format = "#,##0.00"
        yoy_ws.cell(row=row, column=8, value=f"=IF(E{row}=0,\"\",G{row}/ABS(E{row}))").number_format = "0.0%"
        yoy_ws.cell(row=row, column=9, value=(
            f"=IF(AND(ABS(G{row})>{YOY_DOLLAR_CELL},ISNUMBER(H{row}),ABS(H{row})>{YOY_PCT_CELL}),\"FLAG\",\"\")"
        ))
        row += 1
autosize(yoy_ws, {"A": 14, "B": 26, "C": 12, "D": 8, "E": 14, "F": 14, "G": 12, "H": 10, "I": 8})

# ---- MoM_Variance ----

mom_ws = wb.create_sheet("MoM_Variance")
mom_ws.cell(row=1, column=1, value="Month-over-Month Variance - sequential months, 2026 only").font = TITLE_FONT
write_header(mom_ws, 2, ["Account Code", "Account", "Type", "From", "To", "Value (From)", "Value (To)", "$ Var", "% Var", "Flag"])

row = 3
for i in range(len(curr_df)):
    tb_row = TB_DATA_START_ROW + i
    for m1, m2 in zip(MONTHS[:-1], MONTHS[1:]):
        col1, col2 = MONTH_COL[m1], MONTH_COL[m2]
        mom_ws.cell(row=row, column=1, value=f"='TB_2026'!A{tb_row}")
        mom_ws.cell(row=row, column=2, value=f"='TB_2026'!B{tb_row}")
        mom_ws.cell(row=row, column=3, value=f"='TB_2026'!C{tb_row}")
        mom_ws.cell(row=row, column=4, value=m1)
        mom_ws.cell(row=row, column=5, value=m2)
        mom_ws.cell(row=row, column=6, value=f"='TB_2026'!{col1}{tb_row}").number_format = "#,##0.00"
        mom_ws.cell(row=row, column=7, value=f"='TB_2026'!{col2}{tb_row}").number_format = "#,##0.00"
        mom_ws.cell(row=row, column=8, value=f"=G{row}-F{row}").number_format = "#,##0.00"
        mom_ws.cell(row=row, column=9, value=f"=IF(F{row}=0,\"\",H{row}/ABS(F{row}))").number_format = "0.0%"
        mom_ws.cell(row=row, column=10, value=(
            f"=IF(AND(ABS(H{row})>{MOM_DOLLAR_CELL},ISNUMBER(I{row}),ABS(I{row})>{MOM_PCT_CELL}),\"FLAG\",\"\")"
        ))
        row += 1
autosize(mom_ws, {"A": 14, "B": 26, "C": 12, "D": 8, "E": 8, "F": 14, "G": 14, "H": 12, "I": 10, "J": 8})

# ---- Flagged_Summary - every flagged row, tracing back through the variance sheets ----

NOTES = {
    ("COGS - Materials", "YoY"): "Consistent with vendor cost inflation noted in Article 01 - dated and quantified here.",
    ("Marketing & Advertising", "YoY"): "One-month spike, not present in the prior year at all - real campaign or miscoded invoice?",
    ("Marketing & Advertising", "MoM"): "One-month spike, not present in the prior year at all - real campaign or miscoded invoice?",
    ("Accounts Receivable", "YoY"): "Balloons in April 2026 specifically - invisible in a single month, obvious once diffed.",
    ("Accounts Receivable", "MoM"): "Balloons in April 2026 specifically - invisible in a single month, obvious once diffed.",
    ("Professional Fees", "MoM"): "Recurring seasonal pattern - same February spike both years. Not a surprise; flagged here because MoM measures sequential movement, not because it's new.",
}

flag_ws = wb.create_sheet("Flagged_Summary")
flag_ws.cell(row=1, column=1, value="Flagged Accounts - every cell below traces to YoY_Variance / MoM_Variance").font = TITLE_FONT
write_header(flag_ws, 2, ["Type", "Account", "Period", "$ Var", "% Var", "Note"])

row = 3
for i in range(len(yoy_df)):
    r = yoy_df.iloc[i]
    src_row = 3 + i
    if r["flagged"]:
        flag_ws.cell(row=row, column=1, value="YoY")
        flag_ws.cell(row=row, column=2, value=f"=YoY_Variance!B{src_row}")
        flag_ws.cell(row=row, column=3, value=f"=YoY_Variance!D{src_row}")
        flag_ws.cell(row=row, column=4, value=f"=YoY_Variance!G{src_row}").number_format = "#,##0.00"
        flag_ws.cell(row=row, column=5, value=f"=YoY_Variance!H{src_row}").number_format = "0.0%"
        flag_ws.cell(row=row, column=6, value=NOTES.get((r["account"], "YoY"), ""))
        for c in range(1, 7):
            flag_ws.cell(row=row, column=c).fill = FLAG_FILL
        row += 1

for i in range(len(mom_df)):
    r = mom_df.iloc[i]
    src_row = 3 + i
    if r["flagged"]:
        flag_ws.cell(row=row, column=1, value="MoM")
        flag_ws.cell(row=row, column=2, value=f"=MoM_Variance!B{src_row}")
        flag_ws.cell(row=row, column=3, value=f"=MoM_Variance!D{src_row}&\"->\"&MoM_Variance!E{src_row}")
        flag_ws.cell(row=row, column=4, value=f"=MoM_Variance!H{src_row}").number_format = "#,##0.00"
        flag_ws.cell(row=row, column=5, value=f"=MoM_Variance!I{src_row}").number_format = "0.0%"
        note = NOTES.get((r["account"], "MoM"), "")
        if r["recurring"] and not note:
            note = "Recurring seasonal pattern - same swing shows up in the prior year too."
        flag_ws.cell(row=row, column=6, value=note)
        fill = PatternFill("solid", fgColor=WARM_CREAM.lstrip("#")) if r["recurring"] else FLAG_FILL
        for c in range(1, 7):
            flag_ws.cell(row=row, column=c).fill = fill
        row += 1

autosize(flag_ws, {"A": 8, "B": 26, "C": 16, "D": 14, "E": 10, "F": 70})

wb.save(REPORT_PATH)
print(f"[OK] Excel report saved to: {REPORT_PATH}")
