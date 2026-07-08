"""
build_formatting_standard_demo.py - PythonMuse Excel Formatting Standard demo

Builds a small, self-contained teaching workbook that shows the formatting
standard every PythonMuse Excel deliverable follows (header style, number
formats, frozen panes) side by side with the one rule that matters most:
calculated cells must be Excel formulas, never Python-computed values pasted
in as static numbers (CLAUDE.md, Rule 8).

Does not read any real trial balance data - the sample numbers below exist
only to illustrate formatting and are not subject to CLAUDE.md's
source-data rules (Rule 9, Rule 11).

Usage:
    python3 skills/excel-export-standard/scripts/build_formatting_standard_demo.py

Output:
    outputs/excel/formatting_standard_demo.xlsx
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", ".."))
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "outputs", "excel")
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "formatting_standard_demo.xlsx")

# ============================================================
# THE STANDARD - same constants as
# skills/trial-balance-comparison/scripts/generate_visuals.py
# ============================================================

DEEP_NAVY = "002030"
PANEL_NAVY = "0A3040"
GOLD_AMBER = "F8D038"
WRONG_RED = "E8534A"
CORRECT_GREEN = "1E8890"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"

TITLE_FONT = Font(bold=True, size=14, color=DEEP_NAVY)
SUBTITLE_FONT = Font(italic=True, size=10, color=DEEP_NAVY)
HEADER_FILL = PatternFill("solid", fgColor=PANEL_NAVY)
HEADER_FONT = Font(color=WHITE, bold=True)
FLAG_FILL = PatternFill("solid", fgColor=GOLD_AMBER)
WRONG_FILL = PatternFill("solid", fgColor=WRONG_RED)
WRONG_FONT = Font(color=WHITE, bold=True)
CORRECT_FILL = PatternFill("solid", fgColor=CORRECT_GREEN)
CORRECT_FONT = Font(color=WHITE, bold=True)
ALT_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)

CURRENCY_FORMAT = "#,##0.00"
PCT_DETAIL_FORMAT = "0.0%"
PCT_SUMMARY_FORMAT = "0%"


def write_header(ws, row, headers, start_col=1):
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ============================================================
# SHEET 1 - Formatting Standard
# ============================================================

SAMPLE_ACCOUNTS = [
    ("Sales - Demo Widget", 42000.00, 51600.00),
    ("COGS - Demo Materials", 18000.00, 23400.00),
    ("Marketing & Advertising", 1500.00, 9200.00),
]


def build_formatting_standard_sheet(ws):
    ws.cell(row=1, column=1, value="PythonMuse Excel Formatting Standard - Demo").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Header style, number formats, and frozen panes - not a real deliverable").font = SUBTITLE_FONT

    write_header(ws, 4, ["Account", "Q1 Actual", "Q2 Actual", "% Change"])
    ws.freeze_panes = "A5"

    row = 5
    for i, (acct, q1, q2) in enumerate(SAMPLE_ACCOUNTS):
        ws.cell(row=row, column=1, value=acct)
        ws.cell(row=row, column=2, value=q1).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=3, value=q2).number_format = CURRENCY_FORMAT
        pct_cell = ws.cell(row=row, column=4, value=f"=(C{row}-B{row})/ABS(B{row})")
        pct_cell.number_format = PCT_DETAIL_FORMAT
        if i % 2 == 1:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = ALT_FILL
        row += 1

    ws.cell(row=row + 1, column=1, value="Threshold used for flagging (whole-number %):").font = SUBTITLE_FONT
    threshold_cell = ws.cell(row=row + 1, column=4, value=0.15)
    threshold_cell.number_format = PCT_SUMMARY_FORMAT

    autosize(ws, {"A": 26, "B": 14, "C": 14, "D": 12})


# ============================================================
# SHEET 2 - Formula vs. Hardcoded (the rule that matters most)
# ============================================================

VARIANCE_EXAMPLES = [
    ("Sales - Demo Widget", 42000.00, 51600.00),
    ("COGS - Demo Materials", 18000.00, 23400.00),
]


def build_formula_vs_hardcoded_sheet(ws):
    ws.cell(row=1, column=1, value="Formula vs. Hardcoded - CLAUDE.md Rule 8").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="Change a Prior/Current number below, then recalculate (F9). Watch which column updates.",
    ).font = SUBTITLE_FONT

    write_header(ws, 4, ["Account", "Prior", "Current", "$ Var - CORRECT (formula)", "$ Var - WRONG (pasted value)"])
    ws.cell(row=4, column=4).fill = CORRECT_FILL
    ws.cell(row=4, column=4).font = CORRECT_FONT
    ws.cell(row=4, column=5).fill = WRONG_FILL
    ws.cell(row=4, column=5).font = WRONG_FONT
    ws.freeze_panes = "A5"

    row = 5
    for acct, prior, current in VARIANCE_EXAMPLES:
        ws.cell(row=row, column=1, value=acct)
        ws.cell(row=row, column=2, value=prior).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=3, value=current).number_format = CURRENCY_FORMAT

        # CORRECT: a live formula referencing the raw cells on this sheet.
        correct_cell = ws.cell(row=row, column=4, value=f"=C{row}-B{row}")
        correct_cell.number_format = CURRENCY_FORMAT

        # WRONG: the same result, but computed in Python and pasted in as a
        # static number. Looks identical today; silently goes stale the
        # moment Prior/Current changes. This is the anti-pattern, not a
        # style to copy.
        wrong_cell = ws.cell(row=row, column=5, value=round(current - prior, 2))
        wrong_cell.number_format = CURRENCY_FORMAT
        wrong_cell.fill = FLAG_FILL

        row += 1

    ws.cell(row=row + 1, column=1,
             value='CORRECT column: click a cell, formula bar shows something like "=C5-B5" - a reviewer can re-derive it.')
    ws.cell(row=row + 2, column=1,
             value="WRONG column: click a cell, formula bar shows a bare number - a reviewer has no way to tell if it's still accurate.")
    for r in (row + 1, row + 2):
        ws.cell(row=r, column=1).font = SUBTITLE_FONT

    autosize(ws, {"A": 24, "B": 14, "C": 14, "D": 26, "E": 26})


def build():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Formatting Standard"
    build_formatting_standard_sheet(ws1)

    ws2 = wb.create_sheet("Formula vs Hardcoded")
    build_formula_vs_hardcoded_sheet(ws2)

    wb.save(OUTPUT_PATH)
    print(f"[OK] Formatting standard demo saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
